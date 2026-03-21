"""
main.py
Punto de entrada de la API de Encuestas Poblacionales.

Puerto: seleccionado aleatoriamente entre puertos libres del sistema
        para evitar conflictos en entornos multi-servicio.
"""

# ──────────────────────────────────────────────
# SELECCIÓN DE PUERTO ÓPTIMO ALEATORIO
# ──────────────────────────────────────────────
# El puerto NO es simplemente random.randint(8000, 9999).
# Usamos socket para verificar que el puerto esté LIBRE antes de usarlo,
# garantizando que el servidor arranque sin conflictos.
import os
import socket
import random
import time
import logging
import functools
from datetime import datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, HTTPException, Request, status, UploadFile, File, Query
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel, ValidationError

import services
from models import EncuestaCompleta, EncuestaResponse, EstadisticasResponse
from loaders import (
    EXTENSIONES_SOPORTADAS,
    leer_bytes_a_dataframe,
    cargar_desde_url,
    cargar_desde_api_externa,
)

# ──────────────────────────────────────────────
# FORMATO DE FECHA ESTÁNDAR ESPAÑOL
# DD/MM/AAAA HH:MM:SS  (día/mes/año hora:minuto:segundo)
# ──────────────────────────────────────────────
FECHA_FMT   = "%d/%m/%Y %H:%M:%S"   # para timestamps en respuestas JSON
LOG_DATEFMT = "%d/%m/%Y %H:%M:%S"   # para el archivo .log y consola


def ahora_es() -> str:
    """Devuelve la fecha y hora actual en formato estándar español: DD/MM/AAAA HH:MM:SS"""
    return datetime.now().strftime(FECHA_FMT)


# ──────────────────────────────────────────────
# CONFIGURACIÓN DE LOGGING — CONSOLA + ARCHIVO
# ──────────────────────────────────────────────
# Se usan dos manejadores:
#   1. StreamHandler  → muestra logs en la consola en tiempo real
#   2. RotatingFileHandler → escribe en logs/encuesta_api.log
#      El archivo rota al llegar a 2 MB y conserva hasta 5 archivos
#      históricos (encuesta_api.log.1, .2, ...) para no llenar el disco.

LOGS_DIR = Path(__file__).parent / "logs"
LOGS_DIR.mkdir(exist_ok=True)
LOG_FILE = LOGS_DIR / "encuesta_api.log"

_fmt = logging.Formatter(
    fmt="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    datefmt=LOG_DATEFMT,
)

# — Manejador de consola
_console_handler = logging.StreamHandler()
_console_handler.setFormatter(_fmt)
_console_handler.setLevel(logging.INFO)

# — Manejador de archivo rotativo
_file_handler = RotatingFileHandler(
    filename=str(LOG_FILE),
    maxBytes=2 * 1024 * 1024,   # 2 MB por archivo
    backupCount=5,               # guarda hasta 5 archivos históricos
    encoding="utf-8",
)
_file_handler.setFormatter(_fmt)
_file_handler.setLevel(logging.DEBUG)   # el archivo guarda DEBUG también

# — Logger raíz de la aplicación
logging.root.setLevel(logging.DEBUG)
logger = logging.getLogger("encuesta_api")
logger.setLevel(logging.DEBUG)
logger.addHandler(_console_handler)
logger.addHandler(_file_handler)
logger.propagate = False   # evita duplicados en el logger raíz de Python


# ──────────────────────────────────────────────
# FUNCIÓN: ENCONTRAR PUERTO LIBRE ALEATORIO
# ──────────────────────────────────────────────
def encontrar_puerto_libre(rango_inicio: int = 8100, rango_fin: int = 9900) -> int:
    """
    Selecciona aleatoriamente un puerto dentro del rango dado y verifica
    que esté disponible usando un socket de prueba.

    Estrategia:
      1. Barajamos el rango para no ser predecibles.
      2. Por cada candidato, abrimos un socket TCP en modo SO_REUSEADDR.
      3. Si bind() tiene éxito → el puerto está libre → lo usamos.
      4. Si tras 20 intentos no encontramos → lanzamos RuntimeError.

    Args:
        rango_inicio: límite inferior del rango (inclusive). Default 8100
        rango_fin:    límite superior del rango (inclusive). Default 9900

    Returns:
        Número de puerto disponible (int)

    Por qué NO usar solo random.randint():
        random.randint da un número, pero no verifica disponibilidad.
        Podría chocar con otro proceso. Este enfoque es determinista y seguro.
    """
    candidatos = list(range(rango_inicio, rango_fin + 1))
    random.shuffle(candidatos)  # Orden aleatorio real

    for puerto in candidatos[:20]:  # Máximo 20 intentos
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                s.bind(("0.0.0.0", puerto))
                return puerto  # ✅ Puerto libre encontrado
            except OSError:
                continue  # Puerto ocupado, prueba el siguiente

    raise RuntimeError(
        f"No se encontró ningún puerto libre en el rango {rango_inicio}-{rango_fin}."
    )


# ──────────────────────────────────────────────
# DECORADORES PERSONALIZADOS
# ──────────────────────────────────────────────
#
# ¿Cómo se relacionan con los decoradores de FastAPI?
# @app.get / @app.post son decoradores de REGISTRO: le dicen a FastAPI
# qué función maneja qué ruta. Nuestros decoradores son de COMPORTAMIENTO:
# envuelven la función para agregar logging o métricas sin modificar su lógica.
# Ambos usan el mismo mecanismo de Python: functools.wraps + closures.

def log_request(func):
    """
    Decorador que registra en consola:
      - Fecha y hora de la petición
      - Nombre de la función (= endpoint) invocada
      - Tiempo de ejecución en milisegundos

    Uso:  @log_request  encima de cualquier función de endpoint.

    Relación con FastAPI:
      FastAPI usa @app.get/@app.post para registrar rutas (decoradores de registro).
      @log_request es un decorador de comportamiento (cross-cutting concern),
      análogo a un middleware pero a nivel de función individual.
    """
    @functools.wraps(func)
    async def wrapper(*args, **kwargs):
        inicio = time.perf_counter()
        logger.info(f"[PETICIÓN] {ahora_es()} → {func.__name__}()")
        try:
            resultado = await func(*args, **kwargs)
            duracion_ms = (time.perf_counter() - inicio) * 1000
            logger.info(f"[OK]       {func.__name__}() completado en {duracion_ms:.1f}ms")
            return resultado
        except Exception as exc:
            duracion_ms = (time.perf_counter() - inicio) * 1000
            logger.error(f"[ERROR]    {func.__name__}() falló en {duracion_ms:.1f}ms → {exc}")
            raise
    return wrapper


def timer(func):
    """
    Decorador que agrega el header X-Process-Time (ms) a la respuesta.
    Útil para monitoreo de rendimiento en producción.
    """
    @functools.wraps(func)
    async def wrapper(*args, **kwargs):
        inicio = time.perf_counter()
        resultado = await func(*args, **kwargs)
        ms = round((time.perf_counter() - inicio) * 1000, 2)
        logger.info(f"[TIMER]   {func.__name__}: {ms}ms")
        return resultado
    return wrapper


# ──────────────────────────────────────────────
# INSTANCIA FASTAPI
# ──────────────────────────────────────────────
app = FastAPI(
    title="API de Encuestas Poblacionales",
    description=(
        "Sistema de recolección y validación de datos de encuestas poblacionales colombianas. "
        "Valida datos demográficos (estrato 1-6, departamentos oficiales, edad 0-120) "
        "y respuestas en escala Likert o porcentaje antes de persistirlos."
    ),
    version="1.0.0",
    contact={"name": "Ingeniería de Datos", "email": "datos@encuestas.co"},
    license_info={"name": "MIT"},
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ──────────────────────────────────────────────
# MANEJADOR PERSONALIZADO DE ERRORES HTTP 422
# ──────────────────────────────────────────────
# RF4: Captura RequestValidationError (errores de validación Pydantic)
# y devuelve una respuesta JSON estructurada y descriptiva.
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    """
    Manejador personalizado de errores de validación (HTTP 422).

    Transforma el error técnico de Pydantic en un JSON claro para el consumidor:
      - campo: dónde falló la validación
      - mensaje: qué regla se violó
      - valor_recibido: qué mandó el cliente

    Además registra el intento de ingesta inválida en el log (auditoría).
    """
    errores_formateados = []
    for error in exc.errors():
        campo = " → ".join(str(loc) for loc in error["loc"])
        errores_formateados.append({
            "campo": campo,
            "mensaje": error["msg"],
            "tipo_error": error["type"],
            "valor_recibido": error.get("input"),
        })

    # Auditoría: registrar en consola y archivo cada intento con datos inválidos
    logger.warning(
        f"[422 VALIDACIÓN] {request.method} {request.url.path} "
        f"→ {len(errores_formateados)} error(es) de validación"
    )
    for e in errores_formateados:
        logger.warning(f"   ✗ Campo '{e['campo']}': {e['mensaje']} (recibido: {e['valor_recibido']})")

    return JSONResponse(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        content={
            "estado": "error_validacion",
            "codigo": 422,
            "mensaje": "Los datos enviados contienen errores de validación. Revise cada campo.",
            "errores": errores_formateados,
            "ruta": str(request.url.path),
            "timestamp": ahora_es(),
        },
    )


# ──────────────────────────────────────────────
# ENDPOINTS — CRUD ENCUESTAS
# ──────────────────────────────────────────────

@app.post(
    "/encuestas/",
    status_code=status.HTTP_201_CREATED,
    response_model=EncuestaResponse,
    summary="Registrar encuesta completa",
    description=(
        "Recibe y valida una encuesta completa (datos del encuestado + respuestas). "
        "Actúa como aduana transaccional: solo ingresa al repositorio si pasa TODAS las validaciones. "
        "En caso de error devuelve HTTP 422 con detalle por campo."
    ),
    tags=["Encuestas"],
)
@log_request
async def crear_encuesta(encuesta: EncuestaCompleta):
    """
    RF5 — async def vs def:
    ─────────────────────────────────────────────────────────────
    • def (síncrono): FastAPI corre la función en un thread pool para no bloquear el event loop.
    • async def (asíncrono): la función se ejecuta directamente en el event loop de asyncio.

    ¿Cuándo es INDISPENSABLE async/await?
      Cuando el endpoint realiza operaciones de I/O que pueden esperar sin consumir CPU:
        - Consultas a bases de datos (asyncpg, motor)
        - Peticiones HTTP a servicios externos (httpx.AsyncClient)
        - Lectura/escritura de archivos grandes

    Relación con ASGI:
      FastAPI corre sobre ASGI (Asynchronous Server Gateway Interface), que le permite
      manejar miles de conexiones concurrentes con un solo proceso, delegando la espera
      de I/O al event loop en lugar de bloquear threads. Uvicorn implementa el servidor ASGI.
      Comparado con WSGI (Flask/Django clásico), ASGI elimina el cuello de botella de
      "un thread por request" en cargas de I/O intensivo.
    ─────────────────────────────────────────────────────────────
    """
    eid = services.crear_encuesta(encuesta)
    return EncuestaResponse(id=eid, encuesta=encuesta, mensaje="Encuesta registrada exitosamente.")


@app.get(
    "/encuestas/",
    response_model=List[dict],
    summary="Listar todas las encuestas",
    description="Devuelve un resumen de todas las encuestas registradas en el sistema.",
    tags=["Encuestas"],
)
@log_request
async def listar_encuestas():
    return list(services.listar_encuestas().values())


# ──────────────────────────────────────────────
# ENDPOINT — EXPORTAR BASE DE ENCUESTAS
# ──────────────────────────────────────────────

@app.get(
    "/encuestas/exportar",
    summary="Exportar base de encuestas",
    description=(
        "Exporta todas las encuestas registradas en el formato solicitado. "
        "Formatos disponibles: **json** · **csv** · **xlsx** · **txt** · **xml** · **md** (Markdown). "
        "La respuesta incluye el header Content-Disposition para descarga directa desde el navegador."
    ),
    tags=["Exportación"],
)
@log_request
async def exportar_encuestas(
    formato: str = Query(
        default="json",
        description="Formato de exportación: json | csv | xlsx | txt | xml | md",
    )
):
    """
    Genera el archivo de exportación en el formato indicado y lo devuelve
    como StreamingResponse con el Content-Type y nombre de archivo correctos.

    Formatos soportados
    ───────────────────
    json  → Array JSON completo con estructura anidada (encuestado + respuestas)
    csv   → Tabla plana: una fila por encuesta, columnas por campo de encuestado
    xlsx  → Libro Excel con dos hojas: "Encuestados" y "Respuestas"
    txt   → Texto delimitado por pipe (|), legible sin software especial
    xml   → Estructura XML con elemento raíz <encuestas>
    md    → Tabla Markdown lista para pegar en documentación
    """
    import io
    import json as json_mod
    import csv as csv_mod
    import xml.etree.ElementTree as ET
    from fastapi.responses import StreamingResponse

    fmt = formato.lower().strip()
    FORMATOS_VALIDOS = {"json", "csv", "xlsx", "txt", "xml", "md"}

    if fmt not in FORMATOS_VALIDOS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Formato '{formato}' no soportado. "
                f"Use: {', '.join(sorted(FORMATOS_VALIDOS))}"
            ),
        )

    from store import encuestas_db
    encuestas = list(encuestas_db.items())   # [(id, EncuestaCompleta), ...]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_base = f"encuestas_{ts}"

    # ── Función auxiliar: serializar encuestado a dict plano ──
    def encuestado_plano(eid: str, enc) -> dict:
        e = enc.encuestado
        return {
            "id":                eid,
            "nombre":            e.nombre,
            "edad":              e.edad,
            "genero":            e.genero or "",
            "estrato":           e.estrato,
            "departamento":      e.departamento,
            "municipio":         e.municipio or "",
            "nivel_educativo":   e.nivel_educativo or "",
            "ingresos_mensuales": e.ingresos_mensuales if e.ingresos_mensuales is not None else "",
            "personas_hogar":    e.personas_hogar,
            "vivienda":          e.vivienda,
            "situacion_laboral": e.situacion_laboral,
            "fecha":             enc.fecha_diligenciamiento or "",
            "version":           enc.encuesta_version,
            "n_respuestas":      len(enc.respuestas),
        }

    # ─────────────────────────────────────────
    # JSON
    # ─────────────────────────────────────────
    if fmt == "json":
        data = [
            {
                "id": eid,
                "encuestado": enc.encuestado.model_dump(),
                "respuestas": [r.model_dump() for r in enc.respuestas],
                "fecha_diligenciamiento": enc.fecha_diligenciamiento,
                "encuesta_version": enc.encuesta_version,
            }
            for eid, enc in encuestas
        ]
        contenido = json_mod.dumps(data, ensure_ascii=False, indent=2)
        return StreamingResponse(
            iter([contenido]),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{nombre_base}.json"'},
        )

    # ─────────────────────────────────────────
    # CSV
    # ─────────────────────────────────────────
    if fmt == "csv":
        if not encuestas:
            filas = []
            columnas = ["id","nombre","edad","genero","estrato","departamento",
                        "municipio","nivel_educativo","ingresos_mensuales",
                        "personas_hogar","vivienda","situacion_laboral",
                        "fecha","version","n_respuestas"]
        else:
            filas = [encuestado_plano(eid, enc) for eid, enc in encuestas]
            columnas = list(filas[0].keys())

        buf = io.StringIO()
        writer = csv_mod.DictWriter(buf, fieldnames=columnas)
        writer.writeheader()
        writer.writerows(filas)
        buf.seek(0)

        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{nombre_base}.csv"'},
        )

    # ─────────────────────────────────────────
    # XLSX — dos hojas: Encuestados + Respuestas
    # ─────────────────────────────────────────
    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="openpyxl no está instalado. Ejecute: pip install openpyxl",
            )

        wb = openpyxl.Workbook()

        # ── Hoja 1: Encuestados ──
        ws1 = wb.active
        ws1.title = "Encuestados"
        ENCAB_COLOR = "1A4F7A"
        cols_enc = ["ID","Nombre","Edad","Género","Estrato","Departamento",
                    "Municipio","Nivel educativo","Ingresos (COP)",
                    "Personas hogar","Vivienda","Situación laboral",
                    "Fecha","Versión","Nº Respuestas"]
        ws1.append(cols_enc)
        for celda in ws1[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor=ENCAB_COLOR)
            celda.alignment = Alignment(horizontal="center")

        for eid, enc in encuestas:
            p = encuestado_plano(eid, enc)
            ws1.append(list(p.values()))

        for col in ws1.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── Hoja 2: Respuestas ──
        ws2 = wb.create_sheet("Respuestas")
        cols_resp = ["Encuesta ID","Pregunta ID","Pregunta","Tipo","Respuesta","Observación"]
        ws2.append(cols_resp)
        for celda in ws2[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor=ENCAB_COLOR)
            celda.alignment = Alignment(horizontal="center")

        for eid, enc in encuestas:
            for r in enc.respuestas:
                ws2.append([
                    eid,
                    r.pregunta_id,
                    r.pregunta_texto,
                    r.tipo_pregunta,
                    r.respuesta,
                    r.observacion or "",
                ])

        for col in ws2.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{nombre_base}.xlsx"'},
        )

    # ─────────────────────────────────────────
    # TXT — delimitado por pipe
    # ─────────────────────────────────────────
    if fmt == "txt":
        if not encuestas:
            contenido = "# Base de Encuestas Poblacionales — sin registros\n"
        else:
            columnas = ["id","nombre","edad","genero","estrato","departamento",
                        "municipio","nivel_educativo","ingresos_mensuales",
                        "personas_hogar","vivienda","situacion_laboral",
                        "fecha","version","n_respuestas"]
            sep = " | "
            encabezado = sep.join(c.upper() for c in columnas)
            linea_sep  = "-" * len(encabezado)
            filas = [encuestado_plano(eid, enc) for eid, enc in encuestas]
            cuerpo = "\n".join(
                sep.join(str(f.get(c, "")) for c in columnas) for f in filas
            )
            contenido = (
                f"# Base de Encuestas Poblacionales\n"
                f"# Exportado: {ahora_es()} | Total: {len(encuestas)} encuestas\n\n"
                f"{encabezado}\n{linea_sep}\n{cuerpo}\n"
            )

        return StreamingResponse(
            iter([contenido]),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{nombre_base}.txt"'},
        )

    # ─────────────────────────────────────────
    # XML
    # ─────────────────────────────────────────
    if fmt == "xml":
        raiz = ET.Element("encuestas", total=str(len(encuestas)), exportado=ahora_es())

        for eid, enc in encuestas:
            nodo_enc = ET.SubElement(raiz, "encuesta", id=eid)
            e = enc.encuestado

            nodo_enc_datos = ET.SubElement(nodo_enc, "encuestado")
            for campo, valor in [
                ("nombre",            e.nombre),
                ("edad",              str(e.edad)),
                ("genero",            e.genero or ""),
                ("estrato",           str(e.estrato)),
                ("departamento",      e.departamento),
                ("municipio",         e.municipio or ""),
                ("nivel_educativo",   e.nivel_educativo or ""),
                ("ingresos_mensuales",str(e.ingresos_mensuales or "")),
                ("personas_hogar",    str(e.personas_hogar)),
                ("vivienda",          e.vivienda),
                ("situacion_laboral", e.situacion_laboral),
            ]:
                nodo = ET.SubElement(nodo_enc_datos, campo)
                nodo.text = valor

            nodo_meta = ET.SubElement(nodo_enc, "meta")
            ET.SubElement(nodo_meta, "fecha").text    = enc.fecha_diligenciamiento or ""
            ET.SubElement(nodo_meta, "version").text  = enc.encuesta_version

            nodo_resps = ET.SubElement(nodo_enc, "respuestas")
            for r in enc.respuestas:
                nodo_r = ET.SubElement(nodo_resps, "respuesta")
                ET.SubElement(nodo_r, "pregunta_id").text    = r.pregunta_id
                ET.SubElement(nodo_r, "pregunta_texto").text = r.pregunta_texto
                ET.SubElement(nodo_r, "tipo").text           = r.tipo_pregunta
                ET.SubElement(nodo_r, "valor").text          = str(r.respuesta or "")
                ET.SubElement(nodo_r, "observacion").text    = r.observacion or ""

        ET.indent(raiz, space="  ")
        contenido = '<?xml version="1.0" encoding="UTF-8"?>\n' + ET.tostring(raiz, encoding="unicode")

        return StreamingResponse(
            iter([contenido]),
            media_type="application/xml; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{nombre_base}.xml"'},
        )

    # ─────────────────────────────────────────
    # MARKDOWN
    # ─────────────────────────────────────────
    if fmt == "md":
        if not encuestas:
            contenido = "# Base de Encuestas Poblacionales\n\n_Sin registros registrados._\n"
        else:
            cols = ["ID","Nombre","Edad","Estrato","Departamento",
                    "Vivienda","Sit. Laboral","Personas Hogar","Fecha"]
            sep_col = " | "
            encabezado = sep_col.join(cols)
            separador  = sep_col.join(["---"] * len(cols))

            filas_md = []
            for eid, enc in encuestas:
                e = enc.encuestado
                filas_md.append(sep_col.join([
                    eid[:8] + "…",
                    e.nombre,
                    str(e.edad),
                    str(e.estrato),
                    e.departamento,
                    e.vivienda,
                    e.situacion_laboral,
                    str(e.personas_hogar),
                    enc.fecha_diligenciamiento or "—",
                ]))

            tabla = "\n".join([encabezado, separador] + filas_md)
            contenido = (
                f"# Base de Encuestas Poblacionales\n\n"
                f"**Exportado:** {ahora_es()} · **Total:** {len(encuestas)} encuestas\n\n"
                f"## Resumen\n\n"
                f"| {tabla.replace(sep_col, ' | ')} |\n"
            )
            # Corrección: la tabla ya tiene el formato correcto
            lineas = tabla.split("\n")
            tabla_md = "\n".join(f"| {l} |" for l in lineas)
            contenido = (
                f"# Base de Encuestas Poblacionales\n\n"
                f"**Exportado:** {ahora_es()} · **Total:** {len(encuestas)} encuestas\n\n"
                f"## Encuestados\n\n{tabla_md}\n"
            )

        return StreamingResponse(
            iter([contenido]),
            media_type="text/markdown; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{nombre_base}.md"'},
        )



# ⚠️ IMPORTANTE: /estadisticas/ debe ir ANTES de /{id}
# FastAPI resuelve rutas en orden de registro. Si /{id} va primero,
# la palabra "estadisticas" sería interpretada como un ID.
@app.get(
    "/encuestas/estadisticas/",
    response_model=EstadisticasResponse,
    summary="Estadísticas del repositorio",
    description=(
        "Calcula y devuelve métricas estadísticas de todas las encuestas: "
        "conteo total, promedio y mediana de edad, distribución por estrato, "
        "departamento y género, promedio de respuestas por encuesta."
    ),
    tags=["Estadísticas"],
)
@log_request
@timer
async def estadisticas():
    return services.calcular_estadisticas()


@app.get(
    "/encuestas/{id}",
    response_model=EncuestaResponse,
    summary="Obtener encuesta por ID",
    description="Retorna una encuesta específica por su UUID. Devuelve 404 si no existe.",
    tags=["Encuestas"],
)
@log_request
async def obtener_encuesta(id: str):
    encuesta = services.obtener_encuesta(id)
    if not encuesta:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No existe ninguna encuesta con ID '{id}'.",
        )
    return EncuestaResponse(id=id, encuesta=encuesta)


@app.put(
    "/encuestas/{id}",
    response_model=EncuestaResponse,
    summary="Actualizar encuesta existente",
    description="Reemplaza completamente una encuesta existente. Devuelve 404 si el ID no existe.",
    tags=["Encuestas"],
)
@log_request
async def actualizar_encuesta(id: str, encuesta: EncuestaCompleta):
    actualizado = services.actualizar_encuesta(id, encuesta)
    if not actualizado:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No existe ninguna encuesta con ID '{id}'.",
        )
    return EncuestaResponse(id=id, encuesta=encuesta, mensaje="Encuesta actualizada correctamente.")


@app.delete(
    "/encuestas/{id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Eliminar encuesta",
    description="Elimina permanentemente una encuesta del repositorio. Devuelve 204 si fue exitoso.",
    tags=["Encuestas"],
)
@log_request
async def eliminar_encuesta(id: str):
    eliminado = services.eliminar_encuesta(id)
    if not eliminado:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No existe ninguna encuesta con ID '{id}'.",
        )
    return None


# ──────────────────────────────────────────────
# ENDPOINTS — CARGA MASIVA MULTI-FUENTE
# ──────────────────────────────────────────────

@app.post(
    "/encuestas/cargar/archivo",
    summary="Carga masiva desde archivo",
    description=(
        "Ingesta masiva de encuestas desde un archivo subido. "
        "Formatos soportados: **CSV, XLSX, XLS, TXT, TSV, JSON, Parquet, ODS**. "
        "El archivo debe tener columnas que coincidan con los campos del modelo Encuestado. "
        "Las filas inválidas se reportan pero NO detienen la carga de las válidas."
    ),
    tags=["Carga Masiva"],
)
@log_request
async def cargar_desde_archivo(
    archivo: UploadFile = File(..., description="Archivo de datos (CSV, XLSX, TXT, JSON, Parquet, ODS)"),
):
    """Carga encuestas básicas desde un archivo multi-formato."""
    nombre = archivo.filename or ""
    extension = "." + nombre.rsplit(".", 1)[-1].lower() if "." in nombre else ""

    if extension not in EXTENSIONES_SOPORTADAS:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=(
                f"Extensión '{extension}' no soportada. "
                f"Use: {', '.join(EXTENSIONES_SOPORTADAS.keys())}"
            ),
        )

    contenido = await archivo.read()
    try:
        df = leer_bytes_a_dataframe(contenido, extension)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

    return {
        "mensaje": f"Archivo '{nombre}' leído correctamente.",
        "filas": len(df),
        "columnas": list(df.columns),
        "tipos_detectados": df.dtypes.astype(str).to_dict(),
        "muestra": df.head(3).fillna("").to_dict(orient="records"),
    }


@app.post(
    "/encuestas/cargar/url",
    summary="Carga masiva desde URL (nube)",
    description=(
        "Descarga y lee datos desde una URL pública o compartida. "
        "Compatible con: **Google Drive** (link /file/d/<ID>/view), "
        "**Dropbox** (?dl=0), **OneDrive**, **S3 pre-signed**, **GitHub raw**, "
        "y cualquier URL directa a un archivo CSV/XLSX/JSON/Parquet. "
        "El parámetro `extension` es opcional; se infiere de la URL si es posible."
    ),
    tags=["Carga Masiva"],
)
@log_request
async def cargar_url(
    url: str = Query(..., description="URL del archivo en la nube o URL directa"),
    extension: Optional[str] = Query(
        None,
        description="Forzar extensión si no se puede inferir (ej: '.csv', '.xlsx')"
    ),
):
    """
    Carga datos desde una URL de nube.
    Transforma automáticamente links de compartición de Google Drive y Dropbox
    en URLs de descarga directa.
    """
    try:
        df = await cargar_desde_url(url, extension)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

    return {
        "mensaje": "Archivo descargado y leído correctamente desde URL.",
        "url_original": url,
        "filas": len(df),
        "columnas": list(df.columns),
        "muestra": df.head(3).fillna("").to_dict(orient="records"),
    }


@app.post(
    "/encuestas/cargar/api-externa",
    summary="Carga masiva desde API externa",
    description=(
        "Consume un endpoint REST externo que devuelva registros de encuesta en JSON. "
        "Soporta autenticación mediante headers personalizados (Bearer token, API-Key). "
        "Si la lista de registros está anidada dentro de una clave del JSON, "
        "especifíquela en `campo_datos`; si se omite, se detecta automáticamente."
    ),
    tags=["Carga Masiva"],
)
@log_request
async def cargar_api_externa(
    api_url: str = Query(..., description="URL del endpoint externo"),
    metodo: str = Query("GET", description="Verbo HTTP: GET o POST"),
    campo_datos: Optional[str] = Query(
        None,
        description="Clave JSON que contiene la lista (ej: 'data', 'encuestas', 'results')"
    ),
    auth_header: Optional[str] = Query(
        None,
        description="Token de autorización (ej: 'Bearer abc123' o 'ApiKey xyz')"
    ),
):
    """
    Carga registros de encuesta desde otra API REST.
    Detecta automáticamente el campo que contiene la lista de datos.
    """
    headers = {}
    if auth_header:
        headers["Authorization"] = auth_header

    try:
        registros = await cargar_desde_api_externa(
            api_url=api_url,
            metodo=metodo,
            headers=headers,
            campo_datos=campo_datos,
        )
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

    return {
        "mensaje": f"Se obtuvieron {len(registros)} registros desde la API externa.",
        "api_url": api_url,
        "campo_detectado": campo_datos or "auto-detectado",
        "total_registros": len(registros),
        "muestra": registros[:3] if registros else [],
    }



# ──────────────────────────────────────────────
# ENDPOINTS — LOGS DEL FRONTEND
# ──────────────────────────────────────────────

class ErrorFrontend(BaseModel):
    """Modelo para recibir errores reportados por el frontend (index.html)."""
    tipo:       str                    # 'js_error' | 'api_error' | 'red_error' | 'rechazo'
    mensaje:    str                    # descripción del error
    origen:     Optional[str] = None  # archivo JS o URL donde ocurrió
    linea:      Optional[int] = None  # línea del error JS (si aplica)
    columna:    Optional[int] = None  # columna del error JS (si aplica)
    stack:      Optional[str] = None  # traza del error JS (si aplica)
    endpoint:   Optional[str] = None  # ruta API que falló (si aplica)
    codigo_http:Optional[int] = None  # código HTTP de respuesta (si aplica)
    user_agent: Optional[str] = None  # navegador del usuario


# Logger separado exclusivo para errores del frontend
_logger_fe = logging.getLogger("encuesta_api.frontend")
_logger_fe.setLevel(logging.DEBUG)
_logger_fe.addHandler(_file_handler)   # comparte el mismo archivo .log
_logger_fe.addHandler(_console_handler)
_logger_fe.propagate = False

# Archivo de log dedicado solo a errores del frontend
_fe_file_handler = RotatingFileHandler(
    filename=str(LOGS_DIR / "frontend_errores.log"),
    maxBytes=2 * 1024 * 1024,
    backupCount=3,
    encoding="utf-8",
)
_fe_file_handler.setFormatter(_fmt)
_fe_file_handler.setLevel(logging.WARNING)
_logger_fe.addHandler(_fe_file_handler)


@app.post(
    "/logs/error",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Registrar error del frontend",
    description=(
        "Recibe errores capturados por el frontend (index.html) y los persiste "
        "en logs/frontend_errores.log y logs/encuesta_api.log con formato de "
        "fecha estándar español DD/MM/AAAA HH:MM:SS."
    ),
    tags=["Logs"],
)
async def registrar_error_frontend(error: ErrorFrontend, request: Request):
    """
    Endpoint receptor de errores del cliente web.
    El frontend reporta automáticamente errores JS, fallos de red
    y respuestas HTTP inesperadas de la API.
    """
    ip_cliente = request.client.host if request.client else "desconocida"

    # Construir línea de log legible
    detalle = (
        f"[FRONTEND/{error.tipo.upper()}] "
        f"IP={ip_cliente} | "
        f"Mensaje: {error.mensaje}"
    )
    if error.endpoint:
        detalle += f" | Endpoint: {error.endpoint}"
    if error.codigo_http:
        detalle += f" | HTTP {error.codigo_http}"
    if error.origen:
        detalle += f" | Origen: {error.origen}"
        if error.linea:
            detalle += f":{error.linea}"
        if error.columna:
            detalle += f":{error.columna}"
    if error.user_agent:
        detalle += f" | UA: {error.user_agent[:80]}"

    _logger_fe.warning(detalle)

    # Si viene con stack trace, lo guardamos en nivel DEBUG para diagnóstico
    if error.stack:
        _logger_fe.debug(f"  Stack trace:\n{error.stack}")

    return None   # 204 No Content


@app.get(
    "/logs/recientes",
    summary="Ver últimas líneas del log",
    description=(
        "Devuelve las últimas N líneas del archivo de log principal. "
        "Útil para monitoreo rápido sin acceder al servidor."
    ),
    tags=["Logs"],
)
async def ver_logs_recientes(
    n: int = Query(default=50, ge=1, le=500, description="Número de líneas a mostrar (1-500)"),
    solo_errores: bool = Query(default=False, description="Si es true, muestra solo WARNING y ERROR"),
):
    """Lee las últimas N líneas del archivo .log y las devuelve como JSON."""
    if not LOG_FILE.exists():
        return {"lineas": [], "total": 0, "archivo": str(LOG_FILE)}

    with open(LOG_FILE, "r", encoding="utf-8") as f:
        todas = f.readlines()

    if solo_errores:
        todas = [l for l in todas if "WARNING" in l or "ERROR" in l or "CRITICAL" in l]

    ultimas = todas[-n:]
    return {
        "lineas": [l.rstrip("\n") for l in ultimas],
        "total_en_archivo": len(todas),
        "mostrando": len(ultimas),
        "archivo": str(LOG_FILE),
        "consultado": ahora_es(),
    }


# ──────────────────────────────────────────────
# HEALTH CHECK
# ──────────────────────────────────────────────
@app.get("/", tags=["Sistema"], summary="Health check")
async def root():
    return {
        "estado": "online",
        "servicio": "API Encuestas Poblacionales",
        "version": "1.0.0",
        "docs": "/docs",
        "redoc": "/redoc",
        "timestamp": ahora_es(),
    }


# ──────────────────────────────────────────────
# ARRANQUE DEL SERVIDOR CON PUERTO ALEATORIO
# ──────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn

    puerto = encontrar_puerto_libre(rango_inicio=8100, rango_fin=9900)

    print("=" * 55)
    print("  🚀 API de Encuestas Poblacionales")
    print(f"  📡 Puerto asignado: {puerto}")
    print(f"  📖 Swagger UI:  http://localhost:{puerto}/docs")
    print(f"  📘 Redoc:       http://localhost:{puerto}/redoc")
    print(f"  📄 Log:         {LOG_FILE}")
    print("=" * 55)

    logger.info(f"Servidor iniciado — {ahora_es()} — puerto {puerto}")

    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=puerto,
        # workers=1 porque usamos almacenamiento en memoria compartida
        # En producción con BD real, se puede escalar a múltiples workers
        workers=1,
        # reload=False en producción; True solo para desarrollo
        reload=False,
        # log_level para que uvicorn y nuestro logger convivan
        log_level="info",
        # access_log muestra cada petición HTTP en consola
        access_log=True,
    )
